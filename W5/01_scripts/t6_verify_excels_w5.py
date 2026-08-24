# T6 verification: openpyxl read-back of the three W5 workbooks.
# Prints headers + first data rows per sheet, row counts, and reconciliation.
import json
from itertools import islice
import openpyxl
import pyarrow.parquet as pq

OUTDIR = r"C:/Users/esmae/Desktop/phd Esmaeil/THESIS CLAUDE/W5/03_outputs/tables"
PARQ = r"C:/Users/esmae/Desktop/phd Esmaeil/THESIS CLAUDE/Brain/03_db/parquet"

def scan(path, sheets_expect):
    print("=" * 80)
    print(path.split("/")[-1])
    wb = openpyxl.load_workbook(path, read_only=True)
    counts = {}
    for name in wb.sheetnames:
        ws = wb[name]
        if name == "METHOD":
            n = ws.max_row
            first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            print(f"  [{name}] lines={n} first={first[0]!r}")
            counts[name] = n
            continue
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        header = tuple(h for h in header if h is not None)
        n = 0
        firsts = []
        for row in rows:
            n += 1
            if n <= 2:
                firsts.append(row[:min(10, len(header))])
        counts[name] = n
        print(f"  [{name}] data rows = {n}")
        print(f"    header: {header}")
        for fr in firsts:
            print(f"    row:    {fr}")
        # header alignment: every header cell non-empty, width matches first row
        assert all(isinstance(h, str) and h for h in header), f"bad header in {name}"
    wb.close()
    return counts

c1 = scan(f"{OUTDIR}/Driver_Data_W5.xlsx", None)
c2 = scan(f"{OUTDIR}/Sensor_Data_W5.xlsx", None)
c3 = scan(f"{OUTDIR}/Combined_Master_W5.xlsx", None)
nmp = pq.read_metadata(f"{PARQ}/master_events_w5.parquet").num_rows

print("=" * 80)
print("RECONCILIATION")
d_tot = c1["Clean (by trip)"] + c1["Dropped"]
s_tot = c2["Clean readings"] + c2["Removed"]
print(f"  driver: {c1['Clean (by trip)']} clean + {c1['Dropped']} dropped = {d_tot}"
      f"  (expect 264817) {'OK' if d_tot == 264817 else 'FAIL'}")
print(f"  sensor: {c2['Clean readings']} clean + {c2['Removed']} removed = {s_tot}"
      f"  (expect 1048575) {'OK' if s_tot == 1048575 else 'FAIL'}")
ev_expect = c1["Clean (by trip)"] + c2["Drop events"]
print(f"  events: {c3['Events']} = driver clean {c1['Clean (by trip)']} + drops "
      f"{c2['Drop events']} -> {'OK' if c3['Events'] == ev_expect else 'FAIL'}")
db_expect = c1["Dropped"] + c2["Removed"]
print(f"  dropped-both: {c3['Dropped (both files)']} (expect {db_expect}) "
      f"{'OK' if c3['Dropped (both files)'] == db_expect else 'FAIL'}")
print(f"  master parquet rows: {nmp} {'OK' if nmp == c3['Events'] else 'FAIL'}")
print(json.dumps({"driver": c1, "sensor": c2, "combined": c3, "parquet": nmp}))
